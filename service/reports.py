"""Generowanie raportów dla aplikacji serwisowej.

Dwa formaty:

* :func:`generate_quarterly_report_xlsx` — kwartalny raport kosztów per
  maszyna (openpyxl). Wartości tekstowe zaczynające się od ``= + - @`` są
  prefiksowane apostrofem (klasyczna obrona przed CSV / formula injection
  w Excelu — `OWASP Recipe`_).
* :func:`generate_inspection_pdf` — pojedynczy protokół przeglądu (reportlab,
  A4, polskie znaki). Layout celowo prosty żeby drukował się czytelnie z
  domyślnymi marginesami.

.. _OWASP Recipe: https://owasp.org/www-community/attacks/CSV_Injection
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from xml.sax import saxutils

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from core.pdf import font_name, register_pdf_fonts

# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------

# Znaki które Excel/LibreOffice interpretują jako początek formuły. Dowolna
# wartość tekstowa zaczynająca się od jednego z nich może zostać wykonana
# przy otwarciu skoroszytu — prefiksujemy apostrofem żeby zmusić arkusz do
# traktowania całości jako stringa.
_CSV_INJECTION_PREFIXES: tuple[str, ...] = ("=", "+", "-", "@")


def _sanitize(value: str | None) -> str:
    """Defensywnie escapuj wartość przed CSV / formula injection.

    Pusty / ``None`` → pusty string (bezpieczny default dla openpyxl
    ``ws.append``). Niepuste wartości zaczynające się od znaku formuły
    Excela dostają prefix ``'``.
    """
    if not value:
        return ""
    value = str(value)
    if value[:1] in _CSV_INJECTION_PREFIXES:
        return "'" + value
    return value


def _quarter_bounds(year: int, quarter: int) -> tuple[date, date]:
    """Zwróć (start, end) dla danego kwartału kalendarzowego.

    Q1: 01.01-31.03 / Q2: 01.04-30.06 / Q3: 01.07-30.09 / Q4: 01.10-31.12.

    Raises:
        ValueError: ``quarter`` poza zakresem ``1..4``.
    """
    if quarter not in (1, 2, 3, 4):
        raise ValueError(f"Quarter musi być w zakresie 1..4, dostałem {quarter!r}.")
    start_month = (quarter - 1) * 3 + 1
    end_month = start_month + 2
    start_date = date(year, start_month, 1)
    if end_month == 12:
        end_date = date(year, 12, 31)
    else:
        # Pierwszy dzień następnego miesiąca minus jeden — działa też dla
        # kwartałów które kończą się 30 lub 31.
        end_date = date(year, end_month + 1, 1) - date.resolution
    return start_date, end_date


# ----------------------------------------------------------------------------
# XLSX — kwartalny raport kosztów
# ----------------------------------------------------------------------------


_XLSX_HEADERS: tuple[str, ...] = (
    "UID maszyny",
    "Nazwa",
    "Data",
    "Typ",
    "Wykonawca",
    "Opis",
    "Koszt (PLN)",
)


def generate_quarterly_report_xlsx(*, year: int, quarter: int) -> bytes:
    """Wygeneruj XLSX z wpisami serwisowymi za dany kwartał.

    Layout: jeden arkusz nazwany ``Q<n> <year>``, header w pierwszym wierszu
    (bold, biały tekst na niebieskim tle), jeden wiersz na rekord, wiersz
    sumy ("RAZEM:") na końcu. Wszystkie wartości tekstowe są przepuszczone
    przez :func:`_sanitize`.

    Returns:
        XLSX file bytes — gotowe do owinięcia w
        :class:`django.http.HttpResponse` z ``content_type=
        application/vnd.openxmlformats-officedocument.spreadsheetml.sheet``.
    """
    # Import lazy — moduł reports może być importowany przez views.py
    # zanim aplikacja jest fully migrated w testach.
    from service.models import ServiceRecord

    start_date, end_date = _quarter_bounds(year, quarter)

    records = (
        ServiceRecord.objects.select_related("machine")
        .filter(performed_date__gte=start_date, performed_date__lte=end_date)
        .order_by("machine__uid", "performed_date")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Q{quarter} {year}"

    ws.append(list(_XLSX_HEADERS))
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    total = Decimal("0")
    for record in records:
        ws.append(
            [
                _sanitize(record.machine.uid),
                _sanitize(record.machine.name),
                record.performed_date.strftime("%d.%m.%Y"),
                record.get_record_type_display(),
                _sanitize(record.performed_by),
                _sanitize(record.description),
                float(record.cost.amount),
            ]
        )
        total += record.cost.amount

    # Pusty wiersz separujący + RAZEM.
    ws.append([])
    summary_row_idx = ws.max_row + 1
    ws.append(["", "", "", "", "", "RAZEM:", float(total)])
    ws.cell(row=summary_row_idx, column=6).font = Font(bold=True)
    ws.cell(row=summary_row_idx, column=7).font = Font(bold=True)

    # Stała szerokość kolumn — auto-fit w openpyxl jest kosztowny i nieidealny.
    for col_idx in range(1, len(_XLSX_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


_XLSX_HEADERS_FULL: tuple[str, ...] = (
    "UID maszyny",
    "Nazwa",
    "Data",
    "Typ",
    "Wykonawca",
    "Opis",
    "Koszt (PLN)",
    "Następny przegląd",
)


def _write_records_sheet(ws, records, sheet_title: str, headers: tuple[str, ...]) -> Decimal:
    """Pisze pełne pole rekordów do otwartego arkusza openpyxl + zwraca total.

    Layout: header (bold biały na niebieskim tle), wiersze danych, pusty separator,
    wiersz "RAZEM" z sumą kosztów. Stała szerokość kolumn = 18 (auto-fit jest
    drogi, a dla 8 kolumn 18 dobrze pasuje do typowych dat / cen / UID).

    Args:
        ws: openpyxl ``Worksheet`` (już otwarty).
        records: queryset lub iterable ``ServiceRecord``.
        sheet_title: nazwa arkusza (do ``ws.title``).
        headers: lista kolumn — kontroluje który layout (kwartalny vs pełny).

    Returns:
        Suma kosztów (Decimal) — caller może użyć do dalszych obliczeń.
    """
    ws.title = sheet_title
    ws.append(list(headers))
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    total = Decimal("0")
    has_next = "Następny przegląd" in headers
    for record in records:
        row = [
            _sanitize(record.machine.uid),
            _sanitize(record.machine.name),
            record.performed_date.strftime("%d.%m.%Y"),
            record.get_record_type_display(),
            _sanitize(record.performed_by),
            _sanitize(record.description),
            float(record.cost.amount),
        ]
        if has_next:
            row.append(
                record.next_inspection.strftime("%d.%m.%Y") if record.next_inspection else "—"
            )
        ws.append(row)
        total += record.cost.amount

    # Pusty wiersz separujący + RAZEM.
    ws.append([])
    summary_row_idx = ws.max_row + 1
    summary_row = ["", "", "", "", "", "RAZEM:", float(total)]
    if has_next:
        summary_row.append("")
    ws.append(summary_row)
    ws.cell(row=summary_row_idx, column=6).font = Font(bold=True)
    ws.cell(row=summary_row_idx, column=7).font = Font(bold=True)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    return total


def generate_machine_service_xlsx(*, machine) -> bytes:
    """Wygeneruj XLSX z całą historią serwisu jednej maszyny.

    Wszystkie ``ServiceRecord`` przypisane do podanej ``machine`` posortowane
    chronologicznie (od najnowszych). Layout zawiera dodatkową kolumnę
    ``Następny przegląd`` (przydatne na karcie maszyny — operator widzi
    kiedy maszyna ma kolejny obowiązkowy serwis).

    Polskie znaki w nagłówkach i treści są zachowane dzięki natywnej obsłudze
    UTF-8 przez openpyxl (XLSX = ZIP archive z XML w UTF-8).
    """
    from service.models import ServiceRecord

    records = (
        ServiceRecord.objects.select_related("machine")
        .filter(machine=machine)
        .order_by("-performed_date")
    )

    wb = Workbook()
    ws = wb.active
    sheet_title = f"{machine.uid} - Serwis"[:31]  # XLSX limit nazwy arkusza = 31 znaków
    _write_records_sheet(ws, records, sheet_title, _XLSX_HEADERS_FULL)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_all_service_records_xlsx() -> bytes:
    """Wygeneruj XLSX ze WSZYSTKIMI wpisami serwisowymi w bazie.

    Wpisy posortowane: najpierw po UID maszyny rosnąco, potem po dacie
    wykonania malejąco (najnowsze pierwsze w obrębie maszyny). Plus dodatkowa
    kolumna ``Następny przegląd`` jak w wariancie per-maszynowym.

    Używane przez globalny eksport (link "Pobierz wszystkie wpisy" na liście
    serwisów). Operator wyciąga pełną historię floty dla księgowości /
    inspekcji okresowej.
    """
    from service.models import ServiceRecord

    records = ServiceRecord.objects.select_related("machine").order_by(
        "machine__uid", "-performed_date"
    )

    wb = Workbook()
    ws = wb.active
    _write_records_sheet(ws, records, "Pełna historia serwisu", _XLSX_HEADERS_FULL)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_filtered_service_records_xlsx(*, records) -> bytes:
    """Wygeneruj XLSX z PRZEKAZANEJ (już przefiltrowanej) listy wpisów serwisowych.

    Wiersze odpowiadają dokładnie aktywnym filtrom listy serwisowej — eksport
    "Pobierz Excel" zawiera to samo, co widać na ekranie i na wykresie.
    """
    wb = Workbook()
    ws = wb.active
    _write_records_sheet(ws, records, "Historia serwisu (filtr)", _XLSX_HEADERS_FULL)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ----------------------------------------------------------------------------
# PDF — pojedynczy protokół przeglądu
# ----------------------------------------------------------------------------


def generate_inspection_pdf(*, service_record) -> bytes:
    """Wygeneruj protokół przeglądu w PDF (A4, ~1 strona, polskie znaki).

    Layout: tytuł u góry, tabela 11-wierszy z polami protokołu, opis prac
    pod tabelą, miejsce na podpis pod opisem. Numeracja protokołu:
    ``SRV-<6-cyfrowy pk>`` — łatwo cytować w korespondencji.

    Returns:
        PDF file bytes.
    """
    # Bez rejestracji bundled DejaVu Sans reportlab uzywa Helvetica (Latin-1),
    # ktora nie wspiera polskich znakow. Idempotentne — drugie wywolanie noop.
    register_pdf_fonts()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="ProtokolTitle",
        parent=styles["Title"],
        fontName=font_name("bold"),
        fontSize=18,
        alignment=1,  # center
        spaceAfter=20,
    )
    # Override stylow Normal/Bold zeby uzywaly DejaVu (polskie znaki).
    body_style = ParagraphStyle(
        name="ProtokolBody",
        parent=styles["Normal"],
        fontName=font_name(),
        fontSize=11,
        leading=14,
    )

    machine = service_record.machine
    elements = [
        Paragraph("PROTOKÓŁ PRZEGLĄDU TECHNICZNEGO", title_style),
        Spacer(1, 0.5 * cm),
    ]

    table_data = [
        ["Numer protokołu:", f"SRV-{service_record.pk:06d}"],
        ["Data wykonania:", service_record.performed_date.strftime("%d.%m.%Y")],
        ["Typ przeglądu:", service_record.get_record_type_display()],
        ["Maszyna (UID):", machine.uid],
        ["Nazwa maszyny:", machine.name],
        ["Producent:", machine.manufacturer or "-"],
        ["Rok produkcji:", str(machine.build_year) if machine.build_year else "-"],
        ["Numer seryjny:", machine.serial_number or "-"],
        ["Wykonawca:", service_record.performed_by or "-"],
        ["Koszt:", f"{service_record.cost.amount:.2f} {service_record.cost.currency}"],
        [
            "Następny przegląd:",
            service_record.next_inspection.strftime("%d.%m.%Y")
            if service_record.next_inspection
            else "-",
        ],
    ]
    table = Table(table_data, colWidths=[6 * cm, 11 * cm])
    table.setStyle(
        TableStyle(
            [
                ("FONT", (0, 0), (-1, -1), font_name(), 11),
                ("FONT", (0, 0), (0, -1), font_name("bold"), 11),
                ("BOX", (0, 0), (-1, -1), 1, colors.black),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 1 * cm))

    if service_record.description:
        elements.append(
            Paragraph(f'<font name="{font_name("bold")}">Opis prac:</font>', body_style)
        )
        elements.append(Spacer(1, 0.2 * cm))
        # H2 fix: escape user input — reportlab Paragraph parses pseudo-HTML
        # (<a href>, <font>, <img>), więc nieprzetworzony description z user input
        # = phishing/markup injection vector w generowanym PDF.
        elements.append(Paragraph(saxutils.escape(service_record.description), body_style))

    elements.append(Spacer(1, 2 * cm))
    elements.append(
        Paragraph(
            "_________________________<br/>Podpis wykonawcy",
            ParagraphStyle(
                name="Signature",
                parent=body_style,
                alignment=2,  # right
            ),
        )
    )

    doc.build(elements)
    return buffer.getvalue()
