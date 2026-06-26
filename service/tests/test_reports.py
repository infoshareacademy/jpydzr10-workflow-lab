"""Tests for the XLSX / PDF report generation."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook

from service.factories import InspectionFactory, RepairFactory, ServiceRecordFactory
from service.reports import (
    _quarter_bounds,
    _sanitize,
    generate_annual_report_pdf,
    generate_inspection_pdf,
    generate_machine_service_pdf,
    generate_quarterly_report_xlsx,
)

# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------


def test_quarter_bounds_q1():
    start, end = _quarter_bounds(2026, 1)
    assert start == date(2026, 1, 1)
    assert end == date(2026, 3, 31)


def test_quarter_bounds_q2():
    start, end = _quarter_bounds(2026, 2)
    assert start == date(2026, 4, 1)
    assert end == date(2026, 6, 30)


def test_quarter_bounds_q4():
    start, end = _quarter_bounds(2026, 4)
    assert start == date(2026, 10, 1)
    assert end == date(2026, 12, 31)


def test_quarter_bounds_invalid():
    # ``match`` pin'uje na konkretny komunikat — łapie regresję gdyby raise
    # zmieniono na inny ``ValueError`` (np. z innego helpera w kodzie).
    with pytest.raises(ValueError, match=r"Quarter musi być w zakresie 1\.\.4"):
        _quarter_bounds(2026, 5)


def test_sanitize_blocks_formula_prefixes():
    assert _sanitize("=SUM(A1:A10)").startswith("'")
    assert _sanitize("+1+1") == "'+1+1"
    assert _sanitize("-7") == "'-7"
    assert _sanitize("@cmd") == "'@cmd"


def test_sanitize_passes_clean_text():
    assert _sanitize("Jan Kowalski") == "Jan Kowalski"


def test_sanitize_none_or_empty():
    assert _sanitize(None) == ""
    assert _sanitize("") == ""


# ----------------------------------------------------------------------------
# XLSX report
# ----------------------------------------------------------------------------


@pytest.mark.django_db
def test_generate_quarterly_report_returns_bytes(machine):
    ServiceRecordFactory(
        machine=machine,
        performed_date=date(2026, 5, 16),
        cost=Decimal("100.00"),
    )
    data = generate_quarterly_report_xlsx(year=2026, quarter=2)
    assert isinstance(data, bytes)
    assert len(data) > 0


@pytest.mark.django_db
def test_generate_quarterly_report_includes_headers(machine):
    ServiceRecordFactory(machine=machine, performed_date=date(2026, 5, 16))
    payload = generate_quarterly_report_xlsx(year=2026, quarter=2)
    wb = load_workbook(BytesIO(payload), read_only=True)
    ws = wb.active
    first_row = next(ws.iter_rows(values_only=True))
    assert "UID maszyny" in first_row
    # Nagłówek kosztu deklaruje EUR — jedyną walutę po normalizacji (migracja 0004).
    assert "Koszt (EUR)" in first_row


@pytest.mark.django_db
def test_generate_quarterly_report_filters_by_date(machine):
    # Wpis poza Q2 — opis odróżnia go od tego co powinno trafić do raportu.
    ServiceRecordFactory(
        machine=machine,
        performed_date=date(2026, 1, 1),
        cost=Decimal("999.00"),
        description="POZA-ZAKRESEM",
    )
    ServiceRecordFactory(
        machine=machine,
        performed_date=date(2026, 5, 16),
        cost=Decimal("100.00"),
        description="W-ZAKRESIE",
    )
    payload = generate_quarterly_report_xlsx(year=2026, quarter=2)
    wb = load_workbook(BytesIO(payload), read_only=True)
    ws = wb.active
    descriptions = [row[5] for row in ws.iter_rows(values_only=True) if row and row[0]]
    assert "W-ZAKRESIE" in descriptions
    assert "POZA-ZAKRESEM" not in descriptions


@pytest.mark.django_db
def test_generate_quarterly_report_has_total(machine):
    ServiceRecordFactory(machine=machine, performed_date=date(2026, 5, 1), cost=Decimal("100.00"))
    ServiceRecordFactory(machine=machine, performed_date=date(2026, 5, 2), cost=Decimal("250.00"))
    payload = generate_quarterly_report_xlsx(year=2026, quarter=2)
    wb = load_workbook(BytesIO(payload), read_only=True)
    ws = wb.active
    # Last non-empty row should contain "RAZEM:" in column F (6) and 350.0 in column G (7).
    last_row = list(ws.iter_rows(values_only=True))[-1]
    assert "RAZEM:" in last_row
    assert last_row[6] == 350.0


@pytest.mark.django_db
def test_generate_quarterly_report_sanitizes_csv_injection(machine):
    ServiceRecordFactory(
        machine=machine,
        performed_date=date(2026, 5, 16),
        performed_by="=SUM(1+1)",
        cost=Decimal("100.00"),
    )
    payload = generate_quarterly_report_xlsx(year=2026, quarter=2)
    wb = load_workbook(BytesIO(payload), read_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str):
                # Każda formuła musi być sprefiksowana apostrofem przez nasz _sanitize.
                assert not cell.startswith("="), f"Niezabezpieczona formuła: {cell!r}"


@pytest.mark.django_db
def test_generate_quarterly_report_empty_quarter():
    # Brak rekordów — raport nadal valid, headers + RAZEM 0.0.
    payload = generate_quarterly_report_xlsx(year=2026, quarter=1)
    wb = load_workbook(BytesIO(payload), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "UID maszyny"
    last = rows[-1]
    assert "RAZEM:" in last
    assert last[6] == 0.0


# ----------------------------------------------------------------------------
# PDF protokół
# ----------------------------------------------------------------------------


@pytest.mark.django_db
def test_generate_inspection_pdf_returns_bytes(machine):
    record = InspectionFactory(
        machine=machine,
        performed_date=date(2026, 5, 16),
        next_inspection=date(2026, 8, 16),
        performed_by="Jan Kowalski",
        cost=Decimal("750.00"),
        description="Standardowy przegląd kwartalny.",
    )
    payload = generate_inspection_pdf(service_record=record)
    assert isinstance(payload, bytes)
    assert payload.startswith(b"%PDF-"), "Wynik nie wygląda na PDF"
    assert len(payload) > 1000  # Nontrivial size


@pytest.mark.django_db
def test_generate_inspection_pdf_works_for_repair(machine):
    # Naprawa nie ma next_inspection — generator musi obsłużyć "-".
    record = RepairFactory(
        machine=machine,
        performed_date=date(2026, 5, 1),
        next_inspection=None,
    )
    payload = generate_inspection_pdf(service_record=record)
    assert payload.startswith(b"%PDF-")


@pytest.mark.django_db
def test_generate_inspection_pdf_handles_empty_optional_fields(machine):
    # Maszyna bez producenta / sn / build_year — generator nie może crashować.
    machine.manufacturer = ""
    machine.serial_number = ""
    machine.build_year = 0
    machine.save()
    record = InspectionFactory(machine=machine, performed_by="", description="")
    payload = generate_inspection_pdf(service_record=record)
    assert payload.startswith(b"%PDF-")


@pytest.mark.django_db
def test_generate_machine_service_xlsx_returns_xlsx_with_records(machine):
    """Excel per maszyna zawiera tylko wpisy danej maszyny + polskie znaki."""
    from openpyxl import load_workbook

    from service.reports import generate_machine_service_xlsx

    InspectionFactory(
        machine=machine,
        performed_date=date(2026, 5, 16),
        performed_by="Łukasz Żółć",
        description="Smarowanie łożysk + wymiana złącz hydraulicznych.",
        cost=Decimal("450.00"),
    )
    RepairFactory(
        machine=machine,
        performed_date=date(2026, 4, 10),
        description="Naprawa układu sterowania (ząbkowanie).",
    )

    payload = generate_machine_service_xlsx(machine=machine)
    assert isinstance(payload, bytes)
    assert payload[:2] == b"PK"  # XLSX = ZIP

    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws.title.startswith(machine.uid[:31])
    # Header w wierszu 1
    assert ws.cell(1, 1).value == "UID maszyny"
    assert ws.cell(1, 8).value == "Następny przegląd"
    # Wpisy są — sprawdzamy polskie znaki w opisach
    all_text = " ".join(str(c.value or "") for row in ws.iter_rows() for c in row)
    assert "Łukasz" in all_text
    assert "ząbkowanie" in all_text
    assert "RAZEM:" in all_text


@pytest.mark.django_db
def test_generate_all_service_records_xlsx_includes_every_machine(machine):
    """Globalny eksport zawiera wpisy ze wszystkich maszyn."""
    from openpyxl import load_workbook

    from machines.factories import MachineFactory
    from service.reports import generate_all_service_records_xlsx

    second = MachineFactory(uid="K-SECOND", name="Inna koparka")
    InspectionFactory(machine=machine, performed_date=date(2026, 5, 1))
    RepairFactory(machine=second, performed_date=date(2026, 5, 5))

    payload = generate_all_service_records_xlsx()
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    uids = {ws.cell(row, 1).value for row in range(2, ws.max_row + 1)}
    assert machine.uid in uids
    assert "K-SECOND" in uids


@pytest.mark.django_db
def test_generate_inspection_pdf_uses_planer_sans_font_for_polish_chars(machine):
    """Po refaktorze fontów PDF używa PlanerSans (DejaVu) zamiast Helvetica.

    Helvetica jest Latin-1 i nie renderuje polskich znaków. Sprawdzamy że
    po generacji font ``PlanerSans`` jest zarejestrowany w globalnym
    rejestrze reportlab — dzięki temu znaki ą/ę/ł/ó/ś/ż/ź/ć/ń pokazują się
    w gotowym PDFie.
    """
    from reportlab.pdfbase import pdfmetrics

    record = InspectionFactory(
        machine=machine,
        performed_date=date(2026, 5, 16),
        performed_by="Łukasz Żółć",
        description="Naprawa układu hydraulicznego — wymiana łożyska, ząbkowanie.",
        cost=Decimal("900.00"),
    )
    payload = generate_inspection_pdf(service_record=record)
    assert payload.startswith(b"%PDF-")
    # PlanerSans (alias na DejaVu Sans) musi być w rejestrze fontów
    # reportlab po wygenerowaniu PDF — gwarantuje że TableStyle / Paragraph
    # nie spadły do Helvetica fallback.
    registered_fonts = pdfmetrics.getRegisteredFontNames()
    assert "PlanerSans" in registered_fonts, (
        f"PlanerSans nie zarejestrowany, dostępne fonty: {registered_fonts}"
    )


# ----------------------------------------------------------------------------
# Roczny raport PDF + karta serwisowa maszyny PDF (M3 — udostępnione raporty)
# ----------------------------------------------------------------------------


def test_generate_annual_report_pdf_returns_pdf(machine):
    """Roczny raport z wpisami → poprawny, niepusty PDF."""
    InspectionFactory(machine=machine, performed_date=date(2026, 3, 15), cost=Decimal("500.00"))
    RepairFactory(machine=machine, performed_date=date(2026, 7, 1), cost=Decimal("300.00"))
    payload = generate_annual_report_pdf(year=2026)
    assert payload.startswith(b"%PDF-"), "Wynik nie wygląda na PDF"
    assert len(payload) > 1000


def test_generate_annual_report_pdf_empty_year_still_valid(machine):
    """Rok bez wpisów nadal zwraca poprawny PDF (sekcja 'brak wpisów')."""
    payload = generate_annual_report_pdf(year=1999)
    assert payload.startswith(b"%PDF-")


def test_generate_machine_service_pdf_returns_pdf(machine):
    """Karta serwisowa maszyny z historią → poprawny, niepusty PDF."""
    InspectionFactory(machine=machine, performed_date=date(2026, 5, 1), cost=Decimal("120.00"))
    RepairFactory(machine=machine, performed_date=date(2026, 6, 2), cost=Decimal("80.00"))
    payload = generate_machine_service_pdf(machine=machine)
    assert payload.startswith(b"%PDF-")
    assert len(payload) > 1000


def test_generate_machine_service_pdf_no_history_valid(machine):
    """Maszyna bez wpisów serwisowych nadal zwraca poprawny PDF."""
    payload = generate_machine_service_pdf(machine=machine)
    assert payload.startswith(b"%PDF-")


def test_annual_report_pdf_view_downloads(auth_client, machine):
    InspectionFactory(machine=machine, performed_date=date(2026, 3, 1), cost=Decimal("100.00"))
    resp = auth_client.get(reverse("service:report_annual_pdf", kwargs={"year": 2026}))
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/pdf"
    assert "attachment" in resp["Content-Disposition"]


def test_machine_service_pdf_view_downloads(auth_client, machine):
    resp = auth_client.get(reverse("service:report_machine_pdf", kwargs={"uid": machine.uid}))
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/pdf"
    assert "attachment" in resp["Content-Disposition"]


def test_machine_service_pdf_view_404_for_unknown_uid(auth_client):
    resp = auth_client.get(reverse("service:report_machine_pdf", kwargs={"uid": "NOPE-999"}))
    assert resp.status_code == 404


def test_report_pdf_views_require_login(client):
    resp = client.get(reverse("service:report_annual_pdf", kwargs={"year": 2026}))
    assert resp.status_code == 302
