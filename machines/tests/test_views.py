"""View-layer tests for the machines app."""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from django.urls import reverse
from freezegun import freeze_time
from openpyxl import Workbook

from machines.factories import (
    AvailableMachineFactory,
    InServiceMachineFactory,
    MachineFactory,
    OnSiteMachineFactory,
)
from machines.models import Machine


@pytest.mark.django_db
def test_list_view_requires_login(client):
    resp = client.get(reverse("machines:list"))
    assert resp.status_code == 302
    assert "/accounts/login/" in resp.url


@pytest.mark.django_db
def test_list_view_renders(auth_client):
    AvailableMachineFactory.create_batch(3)
    resp = auth_client.get(reverse("machines:list"))
    assert resp.status_code == 200
    assert b"Maszyny" in resp.content


@pytest.mark.django_db
def test_list_view_filters_by_status(auth_client):
    AvailableMachineFactory(uid="A-1")
    OnSiteMachineFactory(uid="O-1")
    resp = auth_client.get(reverse("machines:list"), {"status": Machine.Status.NA_BUDOWIE})
    assert resp.status_code == 200
    assert b"O-1" in resp.content
    assert b"A-1" not in resp.content


@pytest.mark.django_db
def test_list_view_search(auth_client):
    AvailableMachineFactory(uid="FIND-1", name="Specjalna koparka")
    AvailableMachineFactory(uid="OTHER-2", name="Inna")
    resp = auth_client.get(reverse("machines:list"), {"search": "Specjalna"})
    assert b"FIND-1" in resp.content
    assert b"OTHER-2" not in resp.content


@pytest.mark.django_db
@freeze_time("2026-05-16")
def test_list_view_filters_by_inspection_status(auth_client):
    MachineFactory(uid="OVR-1", inspection_date=date(2026, 1, 1))
    MachineFactory(uid="OK-1", inspection_date=date(2026, 12, 1))
    resp = auth_client.get(reverse("machines:list"), {"inspection_status": "overdue"})
    assert b"OVR-1" in resp.content
    assert b"OK-1" not in resp.content


@pytest.mark.django_db
def test_detail_view(auth_client):
    machine = MachineFactory(uid="DET-1", name="Detal test")
    resp = auth_client.get(reverse("machines:detail", kwargs={"uid": machine.uid}))
    assert resp.status_code == 200
    assert b"DET-1" in resp.content
    assert b"Detal test" in resp.content


@pytest.mark.django_db
def test_detail_view_404_for_missing(auth_client):
    resp = auth_client.get(reverse("machines:detail", kwargs={"uid": "NOTEXIST-999"}))
    assert resp.status_code == 404


@pytest.mark.django_db
def test_detail_view_shows_service_records(auth_client):
    """Wave 4 P1: tab "Historia serwisu" pokazuje wpisy z service_records.

    Wcześniej context["service_records"] = [] (placeholder), więc template
    pokazywał "Brak wpisów" mimo że ServiceRecord-y dla maszyny istniały.
    """
    from datetime import date
    from decimal import Decimal

    from service.models import ServiceRecord

    machine = MachineFactory(uid="SVC-DETAIL")
    ServiceRecord.objects.create(
        machine=machine,
        record_type=ServiceRecord.RecordType.PRZEGLAD_KWARTALNY,
        performed_date=date(2026, 1, 15),
        performed_by="Jan Serwisant",
        description="Wymiana oleju + inspekcja hydrauliki",
        cost=Decimal("450.00"),
    )

    resp = auth_client.get(reverse("machines:detail", kwargs={"uid": machine.uid}))
    assert resp.status_code == 200
    records = resp.context["service_records"]
    assert len(records) == 1
    assert records[0].performed_by == "Jan Serwisant"


@pytest.mark.django_db
def test_detail_view_shows_recent_reservations(auth_client):
    """Wave 4 P1: tab "Rezerwacje" pokazuje recent_reservations dla maszyny."""
    from datetime import date, timedelta

    from reservations.factories import ConfirmedReservationFactory

    machine = MachineFactory(uid="RES-DETAIL")
    res = ConfirmedReservationFactory(
        machine=machine,
        start_date=date.today() + timedelta(days=10),
        end_date=date.today() + timedelta(days=15),
        person="Anna Pracownik",
    )

    resp = auth_client.get(reverse("machines:detail", kwargs={"uid": machine.uid}))
    assert resp.status_code == 200
    recent = list(resp.context["recent_reservations"])
    assert len(recent) == 1
    assert recent[0].pk == res.pk


@pytest.mark.django_db
def test_detail_view_history_no_n_plus_1(auth_client, django_assert_max_num_queries):
    """``MachineDetailView`` powinien używać select_related('history_user') przy
    iteracji po historii, żeby renderowanie ``{{ entry.history_user }}`` w
    template nie strzelało N+1 queries (audyt C1-4 P1).

    Tworzymy machine + 5 wpisów history (każdy z innym user — emulacja
    realistycznego scenariusza). Bez select_related — minimum 5 extra queries
    do auth_user (5*query). Z select_related — 1 JOIN.

    Test assertuje max_num_queries=20 jako rozsądny budżet (full render z
    middleware, context processors, auth checks). Główna regresja: BEZ
    fixu N+1 dodaje 5 queries, więc budget 20 łapie szybkie regresje na CI.
    """
    from django.contrib.auth import get_user_model

    user_model = get_user_model()
    machine = MachineFactory(uid="HIST-1", name="Z historią")

    # 5 zmian → 5 wpisów history; każdy z innym user-em (worst case dla N+1).
    for i in range(5):
        u = user_model.objects.create_user(username=f"editor{i}", password="pw")
        machine._history_user = u  # simple-history middleware-style attribution
        machine.notes = f"Edycja {i}"
        machine.save()

    with django_assert_max_num_queries(20):
        resp = auth_client.get(reverse("machines:detail", kwargs={"uid": machine.uid}))
    assert resp.status_code == 200


@pytest.mark.django_db
def test_create_view_requires_permission(auth_client):
    resp = auth_client.get(reverse("machines:create"))
    assert resp.status_code == 403


@pytest.mark.django_db
def test_create_view_post_creates(staff_client):
    payload = {
        "uid": "VW-001",
        "name": "Z formularza",
        "machine_type": Machine.Type.KOPARKA,
        "model": "",
        "capacity": 0,
        "status": Machine.Status.W_MAGAZYNIE,
        "location": "Magazyn",
        "inspection_date": "",
        "manufacturer": "",
        "serial_number": "",
        "build_year": 0,
        "notes": "",
    }
    resp = staff_client.post(reverse("machines:create"), payload)
    assert resp.status_code == 302
    assert Machine.objects.filter(uid="VW-001").exists()


@pytest.mark.django_db
def test_update_view_changes_machine(staff_client):
    machine = MachineFactory(uid="UPD-1", name="Stara nazwa")
    resp = staff_client.post(
        reverse("machines:update", kwargs={"uid": machine.uid}),
        {
            "uid": machine.uid,
            "name": "Nowa nazwa",
            "machine_type": machine.machine_type,
            "model": "",
            "capacity": 0,
            "status": machine.status,
            "location": machine.location,
            "inspection_date": "",
            "manufacturer": "",
            "serial_number": "",
            "build_year": 0,
            "notes": "",
        },
    )
    assert resp.status_code == 302
    machine.refresh_from_db()
    assert machine.name == "Nowa nazwa"


@pytest.mark.django_db
def test_delete_view_removes_machine(staff_client):
    # Cascade collector touches every reverse FK including reservations —
    # skip the check until the F2-B agent ships the reservations migration.
    from django.db import connection

    if "reservations_reservation" not in connection.introspection.table_names():
        pytest.skip("reservations migration not yet shipped (F2-B)")

    machine = MachineFactory(uid="DEL-1")
    resp = staff_client.post(reverse("machines:delete", kwargs={"uid": machine.uid}))
    assert resp.status_code == 302
    assert not Machine.objects.filter(uid="DEL-1").exists()


@pytest.mark.django_db
def test_export_xlsx_returns_attachment(auth_client):
    MachineFactory(uid="EXP-1")
    resp = auth_client.get(reverse("machines:export_xlsx"))
    assert resp.status_code == 200
    assert resp["Content-Type"].startswith("application/vnd.openxmlformats")
    assert "attachment" in resp["Content-Disposition"]


@pytest.mark.django_db
def test_export_xlsx_sanitizes_csv_injection(auth_client):
    """Notatka z formułą Excela jest prefiksowana apostrofem (L2-1 P0 security)."""
    from openpyxl import load_workbook

    MachineFactory(uid="EXP-INJ", notes="=SUM(A1:A100)")
    resp = auth_client.get(reverse("machines:export_xlsx"))
    assert resp.status_code == 200

    workbook = load_workbook(BytesIO(resp.content), read_only=True)
    sheet = workbook.active
    # Znajdź wiersz po UID i wyciągnij notes (ostatnia kolumna w XLSX_COLUMNS).
    notes_value = None
    header = [c.value for c in next(sheet.iter_rows(max_row=1))]
    notes_idx = header.index("notes")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == "EXP-INJ":
            notes_value = row[notes_idx]
            break
    assert notes_value == "'=SUM(A1:A100)", (
        f"Oczekiwano prefiksu apostrofem, otrzymano: {notes_value!r}"
    )


@pytest.mark.django_db
def test_set_service_view_post(staff_client):
    machine = AvailableMachineFactory(uid="SS-1")
    resp = staff_client.post(reverse("machines:set_service", kwargs={"uid": machine.uid}))
    assert resp.status_code == 302
    machine.refresh_from_db()
    assert machine.status == Machine.Status.W_SERWISIE


@pytest.mark.django_db
def test_return_view_post(staff_client):
    machine = InServiceMachineFactory(uid="RT-1")
    resp = staff_client.post(reverse("machines:return", kwargs={"uid": machine.uid}))
    assert resp.status_code == 302
    machine.refresh_from_db()
    assert machine.status == Machine.Status.W_MAGAZYNIE
    assert machine.location == "Magazyn"


# =============================================================================
# RETIRE  (Wave 4 P1 — terminalny status WYCOFANA)
# =============================================================================


@pytest.mark.django_db
def test_retire_view_changes_status(staff_client):
    """MachineRetireView ustawia status WYCOFANA i redirectuje na detail."""
    machine = AvailableMachineFactory(uid="RET-1")
    resp = staff_client.post(
        reverse("machines:retire", kwargs={"uid": machine.uid}),
        data={"reason": "Sprzedana firmie XYZ"},
    )
    assert resp.status_code == 302
    machine.refresh_from_db()
    assert machine.status == Machine.Status.WYCOFANA
    # reason doklejony do notes z prefiksem.
    assert "[WYCOFANA] Sprzedana firmie XYZ" in machine.notes


@pytest.mark.django_db
def test_retire_view_without_reason(staff_client):
    """retire bez reason — status WYCOFANA, notes bez doklejki."""
    machine = AvailableMachineFactory(uid="RET-2")
    original_notes = machine.notes
    resp = staff_client.post(reverse("machines:retire", kwargs={"uid": machine.uid}))
    assert resp.status_code == 302
    machine.refresh_from_db()
    assert machine.status == Machine.Status.WYCOFANA
    assert machine.notes == original_notes  # nie doklejaliśmy nic


@pytest.mark.django_db
def test_retire_view_requires_permission(client, db):
    """User bez ``machines.change_machine`` → 403."""
    from django.contrib.auth import get_user_model

    user_model = get_user_model()
    user = user_model.objects.create_user(username="noperm", password="pw-1234!Tajne")
    client.force_login(user)

    machine = AvailableMachineFactory(uid="RET-NOPERM")
    resp = client.post(reverse("machines:retire", kwargs={"uid": machine.uid}))
    assert resp.status_code == 403
    machine.refresh_from_db()
    # Status nieruszony — guard zadziałał.
    assert machine.status != Machine.Status.WYCOFANA


# =============================================================================
# IMPORT XLSX
# =============================================================================


def _xlsx_with_rows(rows):
    """Build an in-memory XLSX with the import header + given data rows.

    The header order matches the canonical import schema used by
    :class:`MachineImportXlsxView`; only the bare-minimum columns are
    populated so tests stay focused on the view path under test.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["uid", "name", "machine_type", "model", "capacity", "location"])
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    buffer.name = "machines.xlsx"
    return buffer


@pytest.mark.django_db
def test_import_xlsx_creates_machines(staff_client):
    """Two valid rows → two machines persisted."""
    upload = _xlsx_with_rows(
        [
            ["M-T01", "Koparka A", "koparka", "X1", 100, "Magazyn"],
            ["M-T02", "Koparka B", "koparka", "X2", 200, "Magazyn"],
        ]
    )
    resp = staff_client.post(reverse("machines:import_xlsx"), {"file": upload})
    assert resp.status_code in (200, 302)
    assert Machine.objects.filter(uid__in=["M-T01", "M-T02"]).count() == 2


@pytest.mark.django_db
def test_import_xlsx_skips_empty_uid(staff_client):
    """Row with blank ``uid`` is silently skipped, valid row still imports."""
    upload = _xlsx_with_rows(
        [
            ["", "No UID", "koparka", "X", 0, "Magazyn"],
            ["M-T10", "OK", "koparka", "X", 100, "Magazyn"],
        ]
    )
    resp = staff_client.post(reverse("machines:import_xlsx"), {"file": upload})
    assert resp.status_code in (200, 302)
    assert not Machine.objects.filter(name="No UID").exists()
    assert Machine.objects.filter(uid="M-T10").exists()


@pytest.mark.django_db
def test_import_xlsx_handles_malformed_file(staff_client):
    """Plik z poprawnym magic bytes ZIP ale uszkodzoną strukturą — re-render formularza."""
    # Magic bytes ZIP przechodzą walidacje form.clean_file, ale openpyxl nie
    # otworzy uszkodzonego archiwum — caught ``InvalidFileException``/``BadZipFile``
    # i flash error w view.
    fake = BytesIO(b"PK\x03\x04" + b"corrupt-zip-payload")
    fake.name = "bad.xlsx"
    resp = staff_client.post(reverse("machines:import_xlsx"), {"file": fake})
    # View catches openpyxl's exception and re-renders the form (200) instead
    # of leaking a 500.
    assert resp.status_code in (200, 302)
    # No bogus rows were persisted.
    assert not Machine.objects.exists()


@pytest.mark.django_db
def test_import_xlsx_rejects_invalid_magic_bytes_in_view(staff_client):
    """Form-level rejection: plain text z .xlsx ext przerywa zanim dotrze do openpyxl."""
    fake = BytesIO(b"<html>fake</html>")
    fake.name = "fake.xlsx"
    resp = staff_client.post(reverse("machines:import_xlsx"), {"file": fake})
    assert resp.status_code == 200  # re-render z form errors
    assert not Machine.objects.exists()


# =============================================================================
# WAVE 14-F D-3 — inspections_due_modal_view tests
# =============================================================================


@pytest.mark.django_db
def test_inspections_modal_requires_login(client):
    """GET /maszyny/przeglady-w-14d/ bez login → 302 redirect na login."""
    resp = client.get(reverse("machines:inspections_due_modal"))
    assert resp.status_code == 302
    assert "/accounts/login/" in resp.url


@pytest.mark.django_db
def test_inspections_modal_renders_for_logged_user(auth_client):
    """Logged user dostaje 200 + template partial."""
    resp = auth_client.get(reverse("machines:inspections_due_modal"))
    assert resp.status_code == 200
    # Empty-state widoczny gdy brak maszyn (auth_client fixture nie tworzy maszyn).
    assert b"Wszystko aktualne" in resp.content


@pytest.mark.django_db
def test_inspections_modal_lists_overdue_machines(auth_client, machine_factory):
    """Maszyna z inspection_date < today → w sekcji "Przeterminowane"."""
    from datetime import timedelta

    today = date.today()
    machine = machine_factory(
        uid="OVD-INSP-001",
        name="Koparka stara",
        inspection_date=today - timedelta(days=30),
    )

    resp = auth_client.get(reverse("machines:inspections_due_modal"))
    assert resp.status_code == 200
    assert machine in resp.context["machines_overdue"]
    assert b"OVD-INSP-001" in resp.content
    assert b"Przeterminowane" in resp.content


@pytest.mark.django_db
def test_inspections_modal_lists_upcoming_within_14d(auth_client, machine_factory):
    """Maszyna z inspection_date in next 14d → w sekcji "Najbliższe"."""
    from datetime import timedelta

    today = date.today()
    machine = machine_factory(
        uid="UPC-INSP-001",
        name="Walec",
        inspection_date=today + timedelta(days=5),
    )

    resp = auth_client.get(reverse("machines:inspections_due_modal"))
    assert resp.status_code == 200
    assert machine in resp.context["machines_upcoming"]
    assert b"UPC-INSP-001" in resp.content
    assert "najbliższych 14 dniach".encode() in resp.content


@pytest.mark.django_db
def test_inspections_modal_excludes_machines_past_14d(auth_client, machine_factory):
    """Maszyna z inspection_date > today+14 → NIE w upcoming."""
    from datetime import timedelta

    today = date.today()
    machine = machine_factory(
        uid="FAR-INSP-001",
        inspection_date=today + timedelta(days=30),
    )

    resp = auth_client.get(reverse("machines:inspections_due_modal"))
    assert resp.status_code == 200
    assert machine not in resp.context["machines_upcoming"]
    assert machine not in resp.context["machines_overdue"]


@pytest.mark.django_db
def test_inspections_modal_upcoming_limit_20(auth_client, machine_factory):
    """Upcoming queryset limit'owane do max 20 (mobile-friendly)."""
    from datetime import timedelta

    today = date.today()
    # 25 maszyn z przeglądem w next 14d.
    for i in range(25):
        machine_factory(
            uid=f"BULK-{i:03d}",
            inspection_date=today + timedelta(days=1 + (i % 14)),
        )

    resp = auth_client.get(reverse("machines:inspections_due_modal"))
    assert resp.status_code == 200
    assert len(list(resp.context["machines_upcoming"])) == 20


@pytest.mark.django_db
def test_inspections_modal_url_resolves_before_uid_pattern(auth_client):
    """Path "/maszyny/przeglady-w-14d/" matchuje modal view, NIE detail.

    URL conf gotcha: re_path ``(?P<uid>[\\w\\-]+)`` matchuje też
    "przeglady-w-14d" jako uid. Test gwarantuje że nasza ``path()`` jest
    registered PRZED detail (jak w urls.py — defensive regression test).
    """
    resp = auth_client.get(reverse("machines:inspections_due_modal"))
    assert resp.status_code == 200
    assert resp.resolver_match.view_name == "machines:inspections_due_modal"
