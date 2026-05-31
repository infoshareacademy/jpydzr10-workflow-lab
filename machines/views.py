"""Views for the machines app.

Class-based views built on Django's generic CBV stack — each one is a thin
adapter that:

* filters / paginates with help from :class:`MachineFilterForm`,
* delegates writes to the corresponding service in :mod:`machines.services`
  (so the chatbot tool layer in F4 can reuse the same business logic),
* renders one of the Polish templates in ``templates/machines/``.

URL routes are wired in :mod:`machines.urls` (prefix ``/maszyny/``); the
machine UID acts as the slug so URLs read naturally
(``/maszyny/KOP-001/edytuj/``).
"""

from __future__ import annotations

import logging
import zipfile
from datetime import date, timedelta
from io import BytesIO

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils.text import slugify
from django.utils.translation import gettext as _
from django.views import View
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    FormView,
    ListView,
    UpdateView,
)
from openpyxl.utils.exceptions import InvalidFileException

from core.pagination import PerPageMixin

from .forms import MachineFilterForm, MachineForm, MachineImportXlsxForm
from .models import Machine
from .services import (
    close_repair,
    create_machine,
    retire_machine,
    return_machine_to_warehouse,
    set_machine_to_service,
    update_machine,
)

logger = logging.getLogger("machines")


# =============================================================================
# LIST
# =============================================================================


class MachineListView(PerPageMixin, LoginRequiredMixin, ListView):
    """Paginated list of machines with sidebar filters.

    PerPageMixin: ?per_page=N (whitelist 10/20/50/100/500/5000), default 100.
    """

    model = Machine
    template_name = "machines/list.html"
    context_object_name = "machines"

    def get_queryset(self):
        queryset = Machine.objects.all()
        self.filter_form = MachineFilterForm(self.request.GET or None)
        if self.filter_form.is_valid():
            data = self.filter_form.cleaned_data
            if data.get("search"):
                term = data["search"]
                queryset = queryset.filter(Q(uid__icontains=term) | Q(name__icontains=term))
            if data.get("status"):
                queryset = queryset.filter(status=data["status"])
            if data.get("machine_type"):
                queryset = queryset.filter(machine_type=data["machine_type"])
            if data.get("inspection_status"):
                queryset = _apply_inspection_filter(queryset, data["inspection_status"])
            if data.get("is_reservable") == "yes":
                queryset = queryset.filter(is_reservable=True)
            elif data.get("is_reservable") == "no":
                queryset = queryset.filter(is_reservable=False)
        elif self.request.GET:
            # User wpisał ?status=nonsense albo podobne — silently rezygnujemy
            # z filtrowania, ale informujemy o tym przez messages.warning.
            invalid_fields = ", ".join(self.filter_form.errors.keys())
            messages.warning(
                self.request,
                _(
                    "Pominięto nieprawidłowe wartości filtrów: %(fields)s. "
                    "Wyświetlono wszystkie maszyny."
                )
                % {"fields": invalid_fields},
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter_form"] = self.filter_form
        context["total_count"] = Machine.objects.count()
        return context


def _apply_inspection_filter(queryset, bucket: str):
    """Translate the four ``inspection_status`` buckets into ORM filters."""
    today = date.today()
    warning_until = today + timedelta(days=14)
    if bucket == "ok":
        return queryset.filter(inspection_date__gt=warning_until)
    if bucket == "warning":
        return queryset.filter(inspection_date__gte=today, inspection_date__lte=warning_until)
    if bucket == "overdue":
        return queryset.filter(inspection_date__lt=today)
    if bucket == "unknown":
        return queryset.filter(inspection_date__isnull=True)
    return queryset


# =============================================================================
# DETAIL
# =============================================================================


class MachineDetailView(LoginRequiredMixin, DetailView):
    """Single-machine page z trzema zakładkami: Dane, Serwis, Rezerwacje, Historia.

    Wave 4 P1 fix: service_records i recent_reservations były wcześniej
    placeholdered jako puste listy mimo że F3 (service) i F2 (reservations)
    agents zostały dawno doszyte — template pokazywał "Brak wpisów" mimo
    że dane istniały. Teraz ładujemy faktyczne queries z prefetchem.
    """

    model = Machine
    template_name = "machines/detail.html"
    context_object_name = "machine"
    slug_field = "uid"
    slug_url_kwarg = "uid"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        machine = self.object

        # Wave 4 P1: service_records — ServiceRecord.machine FK z
        # related_name="service_records". Limit 20 ostatnich (zakładka
        # ma link "Wszystkie wpisy →" prowadzący na pełną listę).
        # select_related na performed_by nie potrzebne — to CharField, nie FK.
        context["service_records"] = machine.service_records.order_by("-performed_date")[:20]

        # Wave 4 P1: recent_reservations — Reservation.machine FK z
        # related_name="reservations". Limit 10, najnowsze ostatnich.
        # Lazy import — unikamy circular import (reservations importuje machines).
        # Wykluczamy ANULOWANA — to były rezerwacje "odwołane" przed startem,
        # nie reprezentują pracy maszyny. Bez exclude lista 10 zostawała by
        # zdominowana przez anulowane, ważne historyczne by wypadły poza top-10.
        from reservations.models import Reservation

        context["recent_reservations"] = (
            Reservation.objects.filter(machine=machine)
            .exclude(status=Reservation.Status.ANULOWANA)
            .select_related("site")
            .order_by("-start_date")[:10]
        )

        # ``select_related("history_user")`` — bez tego template renderujący
        # ``{{ entry.history_user }}`` strzela N+1 query (5 extra queries dla
        # slice [:5]). Z select_related → 1 JOIN do auth_user (C1-4 P1 audyt).
        context["history"] = machine.history.select_related("history_user")[:5]
        return context


# =============================================================================
# CREATE / UPDATE / DELETE
# =============================================================================


class MachineCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Operator-only form for adding a new machine."""

    model = Machine
    form_class = MachineForm
    template_name = "machines/form.html"
    permission_required = "machines.add_machine"
    raise_exception = True  # P1 fix: 403 zamiast login redirect dla authed-bez-perm

    def form_valid(self, form):
        try:
            machine = create_machine(**form.cleaned_data)
        except ValidationError as exc:
            form.add_error(None, exc)
            return self.form_invalid(form)
        messages.success(self.request, _("Maszyna %(uid)s została dodana.") % {"uid": machine.uid})
        return redirect("machines:detail", uid=machine.uid)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form_title"] = _("Dodaj maszynę")
        context["submit_label"] = _("Dodaj maszynę")
        return context


class MachineUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Operator-only form for editing an existing machine."""

    model = Machine
    form_class = MachineForm
    template_name = "machines/form.html"
    slug_field = "uid"
    slug_url_kwarg = "uid"
    permission_required = "machines.change_machine"
    raise_exception = True  # P1 fix: 403 zamiast login redirect dla authed-bez-perm

    def form_valid(self, form):
        try:
            machine, warnings = update_machine(self.object, **form.cleaned_data)
        except ValidationError as exc:
            form.add_error(None, exc)
            return self.form_invalid(form)
        messages.success(
            self.request,
            _("Maszyna %(uid)s została zaktualizowana.") % {"uid": machine.uid},
        )
        for warning in warnings:
            messages.warning(self.request, warning)
        return redirect("machines:detail", uid=machine.uid)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form_title"] = _("Edytuj maszynę %(uid)s") % {"uid": self.object.uid}
        context["submit_label"] = _("Zapisz zmiany")
        return context


class MachineDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Hard delete — guarded by a confirmation template."""

    model = Machine
    template_name = "machines/delete_confirm.html"
    slug_field = "uid"
    slug_url_kwarg = "uid"
    permission_required = "machines.delete_machine"
    raise_exception = True  # P1 fix: 403 zamiast login redirect dla authed-bez-perm
    success_url = reverse_lazy("machines:list")

    def form_valid(self, form):
        uid = self.object.uid
        response = super().form_valid(form)
        messages.success(self.request, _("Maszyna %(uid)s została usunięta.") % {"uid": uid})
        return response


# =============================================================================
# IMPORT / EXPORT XLSX
# =============================================================================


XLSX_COLUMNS: tuple[str, ...] = (
    "uid",
    "name",
    "machine_type",
    "model",
    "capacity",
    "status",
    "location",
    "inspection_date",
    "manufacturer",
    "serial_number",
    "build_year",
    "notes",
)


class MachineImportXlsxView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """Bulk upload — one row per machine, columns match :data:`XLSX_COLUMNS`."""

    template_name = "machines/import.html"
    form_class = MachineImportXlsxForm
    permission_required = "machines.add_machine"
    raise_exception = True  # P1 fix: 403 zamiast login redirect dla authed-bez-perm
    success_url = reverse_lazy("machines:list")

    def form_valid(self, form):
        upload = form.cleaned_data["file"]
        try:
            workbook = openpyxl.load_workbook(upload, data_only=True, read_only=True)
        except (InvalidFileException, zipfile.BadZipFile, KeyError, OSError) as exc:
            # Logujemy z stack trace (logger.exception) — bare ``except Exception``
            # ukrywało prawdziwy root cause crashy w produkcji. Wąska lista klas
            # pokrywa typowe scenariusze: zły format, uszkodzony ZIP (.xlsx to ZIP),
            # brakujące struktury wewnątrz pliku, błędy IO.
            # M2 fix: sanitize filename przed log line — multipart filename
            # może zawierać \r\n które forguje log entries (CRLF injection).
            safe_name = upload.name.replace("\r", "").replace("\n", "")
            logger.exception("XLSX import failed: %s", safe_name)
            messages.error(
                self.request,
                _("Nie udało się odczytać pliku XLSX: %(error)s") % {"error": exc},
            )
            return self.form_invalid(form)

        sheet = workbook.active
        header = [str(c.value).strip() if c.value else "" for c in next(sheet.iter_rows(max_row=1))]
        created, skipped, errors = 0, 0, []

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = dict(zip(header, row, strict=False))
            uid = (
                (row_dict.get("uid") or "").strip()
                if isinstance(row_dict.get("uid"), str)
                else row_dict.get("uid")
            )
            if not uid:
                skipped += 1
                continue
            try:
                create_machine(
                    uid=str(uid),
                    name=str(row_dict.get("name") or ""),
                    machine_type=str(row_dict.get("machine_type") or Machine.Type.INNE),
                    model=str(row_dict.get("model") or ""),
                    capacity=int(row_dict.get("capacity") or 0),
                    status=str(row_dict.get("status") or Machine.Status.W_MAGAZYNIE),
                    location=str(row_dict.get("location") or "Magazyn"),
                    inspection_date=row_dict.get("inspection_date") or None,
                    manufacturer=str(row_dict.get("manufacturer") or ""),
                    serial_number=str(row_dict.get("serial_number") or ""),
                    build_year=int(row_dict.get("build_year") or 0),
                    notes=str(row_dict.get("notes") or ""),
                )
                created += 1
            except (ValidationError, ValueError, TypeError) as exc:
                errors.append(
                    _("Wiersz %(row)d (%(uid)s): %(error)s")
                    % {"row": row_index, "uid": uid, "error": exc}
                )

        if created:
            messages.success(
                self.request, _("Zaimportowano %(count)d maszyn.") % {"count": created}
            )
        if skipped:
            messages.warning(
                self.request,
                _("Pominięto %(count)d pustych wierszy.") % {"count": skipped},
            )
        for err in errors[:10]:
            messages.error(self.request, err)
        if len(errors) > 10:
            messages.error(
                self.request,
                _("...oraz %(count)d dalszych błędów.") % {"count": len(errors) - 10},
            )
        return super().form_valid(form)


# Znaki które Excel/LibreOffice interpretują jako początek formuły. Dowolna
# wartość tekstowa zaczynająca się od jednego z nich może zostać wykonana
# przy otwarciu skoroszytu — prefiksujemy apostrofem żeby zmusić arkusz do
# traktowania całości jako stringa. Mirror analogicznego helpera w
# ``service.reports._sanitize`` (potencjalny refactor do core/security.py, P1).
_XLSX_FORMULA_PREFIXES: tuple[str, ...] = ("=", "+", "-", "@")


def _sanitize_xlsx_cell(value):
    """Defensywnie escapuj wartość przed CSV / formula injection w XLSX.

    Chroni przed atakami typu ``=cmd|'/c calc'!A0`` — pole tekstowe maszyny
    z formułą Excela mogłoby zostać wykonane przy otwarciu eksportu przez
    audytora (L2-1 P0 fix). Wartości nieznakowe zwracane bez zmian (liczby,
    daty trzymają typ).
    """
    if not isinstance(value, str):
        return value
    if value and value[0] in _XLSX_FORMULA_PREFIXES:
        return "'" + value
    return value


class MachineExportXlsxView(LoginRequiredMixin, View):
    """Streaming XLSX download of the full machine inventory."""

    def get(self, request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Maszyny"
        sheet.append(list(XLSX_COLUMNS))

        for machine in Machine.objects.all().order_by("uid"):
            sheet.append(
                [
                    _sanitize_xlsx_cell(machine.uid),
                    _sanitize_xlsx_cell(machine.name),
                    _sanitize_xlsx_cell(machine.machine_type),
                    _sanitize_xlsx_cell(machine.model),
                    machine.capacity,
                    _sanitize_xlsx_cell(machine.status),
                    _sanitize_xlsx_cell(machine.location),
                    machine.inspection_date.isoformat() if machine.inspection_date else "",
                    _sanitize_xlsx_cell(machine.manufacturer),
                    _sanitize_xlsx_cell(machine.serial_number),
                    machine.build_year,
                    _sanitize_xlsx_cell(machine.notes or ""),
                ]
            )

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        filename = f"maszyny-{slugify(date.today().isoformat())}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# =============================================================================
# STATUS ACTIONS (POST only)
# =============================================================================


class MachineSetServiceView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """``POST`` endpoint — call :func:`set_machine_to_service`."""

    permission_required = "machines.change_machine"
    raise_exception = True  # P1 fix: 403 zamiast login redirect dla authed-bez-perm

    def post(self, request, uid: str):
        machine = get_object_or_404(Machine, uid=uid)
        try:
            set_machine_to_service(machine)
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        else:
            messages.success(
                request,
                _("Maszyna %(uid)s została wysłana do serwisu.") % {"uid": machine.uid},
            )
        return redirect("machines:detail", uid=machine.uid)


class MachineReturnView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """``POST`` endpoint — call :func:`return_machine_to_warehouse`."""

    permission_required = "machines.change_machine"
    raise_exception = True  # P1 fix: 403 zamiast login redirect dla authed-bez-perm

    def post(self, request, uid: str):
        machine = get_object_or_404(Machine, uid=uid)
        result = return_machine_to_warehouse(machine)
        closed = result["closed"]
        if closed:
            messages.success(
                request,
                _(
                    "Maszyna %(uid)s została zwrócona do magazynu "
                    "(zamknięto %(closed)d aktywnych rezerwacji)."
                )
                % {"uid": machine.uid, "closed": closed},
            )
        else:
            messages.success(
                request,
                _("Maszyna %(uid)s została zwrócona do magazynu.") % {"uid": machine.uid},
            )
        return redirect("machines:detail", uid=machine.uid)


class MachineCloseRepairView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """``POST`` endpoint — przełącza maszynę z ``W serwisie`` z powrotem do magazynu."""

    permission_required = "machines.change_machine"
    raise_exception = True  # P1 fix: 403 zamiast login redirect dla authed-bez-perm

    def post(self, request, uid: str):
        machine = get_object_or_404(Machine, uid=uid)
        try:
            close_repair(machine)
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        else:
            messages.success(
                request,
                _("Naprawa zakończona — maszyna %(uid)s dostępna w magazynie.")
                % {"uid": machine.uid},
            )
        return redirect("machines:detail", uid=machine.uid)


class MachineRetireView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """``POST`` endpoint — trwale wycofuje maszynę z floty (status=WYCOFANA).

    Wave 4 P1: ``retire_machine`` service istniał i był testowany, ale BRAK
    było URL/view/button — orphan function. Teraz operator może wycofać
    maszynę z UI (sprzedana / złomowana / utracona). Akcja nieodwracalna
    (terminalny status), więc template wymaga potwierdzenia + powodu.

    Opcjonalne pole ``reason`` doklejane do ``machine.notes`` jako
    ``[WYCOFANA] <reason>`` (pełna historia w simple_history).
    """

    permission_required = "machines.change_machine"
    raise_exception = True  # 403 dla authed bez perm

    def post(self, request, uid: str):
        machine = get_object_or_404(Machine, uid=uid)
        reason = (request.POST.get("reason") or "").strip()
        try:
            retire_machine(machine, reason=reason)
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        else:
            messages.success(
                request,
                _("Maszyna %(uid)s została wycofana z floty.") % {"uid": machine.uid},
            )
        return redirect("machines:detail", uid=machine.uid)


# =============================================================================
# WAVE 14-F D-3 — Inspections due modal (HTMX partial)
# =============================================================================


@login_required
def inspections_due_modal_view(request):
    """Wave 14-F D-3: HTMX partial — lista maszyn z przeglądem 14d + overdue.

    Audyt Wave 14-E D-3: KPI card "Przeglądy w 14 dniach" na dashboardzie
    pokazuje liczbę (np. "3 maszyny przegląd 14d"), ale operator nie wie
    KTÓRE konkretnie maszyny mają zaraz przegląd i KIEDY. Sebastian
    walkthrough 17 maja: "kliknij → lista maszyn z datami".

    View jest wywoływany przez HTMX (``hx-get`` na KPI card) i swap'iuje
    HTML do ``<div id="inspections-modal-content">`` w home.html. Alpine
    listener ``@open-modal.window`` odpala visibility modala (event
    dispatch z hx-on::after-request).

    Querysets:

    * ``overdue`` — :meth:`MachineManager.overdue_inspection` (inspection_date
      strictly < today), ALL (no limit — operator musi widzieć wszystkie
      blokery legalnej pracy maszyny).
    * ``upcoming`` — :meth:`MachineManager.upcoming_inspection` (next 14 days)
      max [:20] — dłuższe listy nieczytelne na mobile + i tak chcemy je
      filtrować przez /maszyny/?inspection_status=warning.

    Permission: ``@login_required`` — inspection data nie zawiera PII
    (tylko UID + nazwa maszyny + data), ale zostawiamy spójność z resztą
    widoków dashboardowych.
    """
    today = date.today()
    overdue_qs = Machine.objects.overdue_inspection(today=today).order_by("inspection_date")
    upcoming_qs = Machine.objects.upcoming_inspection(days=14, today=today).order_by(
        "inspection_date"
    )[:20]
    return render(
        request,
        "machines/_inspections_due_modal.html",
        {
            "machines_overdue": overdue_qs,
            "machines_upcoming": upcoming_qs,
            "today": today,
        },
    )
